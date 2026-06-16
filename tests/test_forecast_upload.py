"""Unit tests for the bulk-upload logic in 销售预测/预测录入.

The page module has a non-importable name (starts with a digit, Chinese
chars) and runs Streamlit at import time, so we load it by path with
importlib. The Streamlit calls at module level only emit "missing
ScriptRunContext" warnings outside a real runtime — they don't raise.
"""
import io
import importlib.util
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# ── Load the page module by path ──────────────────────────────────────────
_PAGE = (Path(__file__).resolve().parent.parent
         / "app" / "pages" / "02_销售预测.py")
_spec = importlib.util.spec_from_file_location("forecast_page", _PAGE)
forecast_page = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(forecast_page)

make_template = forecast_page._make_excel_template
parse_and_validate = forecast_page.parse_and_validate


# Chinese headers exactly as produced by the template (keys of _COL_MAP)
H = {
    "shipper_id": "货主编号*",
    "customer_id": "客户编号*",
    "destination": "目的地代码*",
    "sku_id": "SKU编号*",
    "quantity_pallets": "预测数量（托盘）*",
    "required_date": "要求交付日期*",
}


def _df(rows):
    """Build a raw upload DataFrame with the template's Chinese headers."""
    return pd.DataFrame([{H[k]: v for k, v in row.items()} for row in rows])


def _valid_row(**overrides):
    row = {
        "shipper_id": "SH001",
        "customer_id": "CUST001",
        "destination": "CC",
        "sku_id": "SKU001",
        "quantity_pallets": 15,
        "required_date": "2026-06-23",  # a Tuesday
    }
    row.update(overrides)
    return row


# ── Template generation ───────────────────────────────────────────────────
def test_template_has_two_sheets_and_headers():
    data = make_template()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["预测数据", "填写说明"]

    ws = wb["预测数据"]
    headers = [ws.cell(1, c).value for c in range(1, 7)]
    assert headers == [
        "货主编号*", "客户编号*", "目的地代码*",
        "SKU编号*", "预测数量（托盘）*", "要求交付日期*",
    ]


def test_template_all_columns_marked_required():
    """All six headers carry the '*' required marker."""
    wb = openpyxl.load_workbook(io.BytesIO(make_template()))
    ws = wb["预测数据"]
    headers = [ws.cell(1, c).value for c in range(1, 7)]
    assert all(h.endswith("*") for h in headers)


# ── Valid data ────────────────────────────────────────────────────────────
def test_valid_data_passes_and_normalizes():
    df, errors = parse_and_validate(_df([_valid_row(), _valid_row(customer_id="CUST002")]))
    assert errors == []
    assert len(df) == 2
    # date snapped from Tuesday 2026-06-23 to Monday 2026-06-22
    assert df["required_date"].iloc[0] == date(2026, 6, 22)
    # quantity is int
    assert df["quantity_pallets"].iloc[0] == 15
    assert str(df["quantity_pallets"].dtype).startswith("int")


def test_date_already_monday_unchanged():
    df, errors = parse_and_validate(_df([_valid_row(required_date="2026-06-22")]))
    assert errors == []
    assert df["required_date"].iloc[0] == date(2026, 6, 22)


def test_blank_rows_are_dropped():
    rows = _df([_valid_row()])
    # append a fully-empty row
    rows.loc[len(rows)] = [None] * len(rows.columns)
    df, errors = parse_and_validate(rows)
    assert errors == []
    assert len(df) == 1


# ── Missing column ────────────────────────────────────────────────────────
def test_missing_required_column():
    raw = _df([_valid_row()]).drop(columns=[H["customer_id"]])
    df, errors = parse_and_validate(raw)
    assert df is None
    assert any("客户编号" in e and "缺少必填列" in e for e in errors)


# ── Empty values ──────────────────────────────────────────────────────────
def test_empty_value_in_required_column():
    df, errors = parse_and_validate(_df([_valid_row(quantity_pallets=None)]))
    assert df is None
    assert any("预测数量" in e and "空值" in e for e in errors)


def test_empty_value_reports_row_number():
    # row 0 valid, row 1 has empty shipper → reported as Excel row 3
    df, errors = parse_and_validate(_df([_valid_row(), _valid_row(shipper_id=None)]))
    assert df is None
    assert any("货主编号" in e and "3" in e for e in errors)


# ── Invalid enum values ───────────────────────────────────────────────────
@pytest.mark.parametrize("field,bad,needle", [
    ("customer_id", "CUST999", "无效客户编号"),
    ("destination", "XX", "无效目的地代码"),
    ("sku_id", "SKU999", "无效SKU编号"),
    ("shipper_id", "SH999", "无效货主编号"),
])
def test_invalid_enum_value(field, bad, needle):
    df, errors = parse_and_validate(_df([_valid_row(**{field: bad})]))
    assert df is None
    assert any(needle in e for e in errors)


def test_multiple_errors_accumulate():
    df, errors = parse_and_validate(_df([_valid_row(customer_id="CUST999", destination="XX")]))
    assert df is None
    assert len(errors) >= 2
