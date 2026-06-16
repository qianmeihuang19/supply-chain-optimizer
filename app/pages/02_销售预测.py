"""Page 2: Sales Forecast Management."""
import io
import streamlit as st
import pandas as pd
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

DEST_OPTIONS = {"CC-长春": "CC", "DL-大连": "DL", "TJ-天津": "TJ"}
CUSTOMERS = ["CUST001", "CUST002", "CUST003"]
SKU_OPTIONS = ["SKU001"]  # single SKU demo; details managed in 系统管理

# Enum allowed values, keyed by English column key
_ENUMS = {
    "shipper_id": ["SH001", "SH002"],
    "customer_id": ["CUST001", "CUST002", "CUST003"],
    "destination": ["CC", "DL", "TJ"],
    "sku_id": ["SKU001"],
}

# ── Column specs ──────────────────────────────────────────────────────────
# Each column: header (Chinese, * = required), key (English), width, example,
# kind ("enum"|"int"|"date"|"text"), required, label (for messages),
# allowed/fmt (填写说明 text), snap (date→Monday).
FORECAST_COLUMNS = [
    {"header": "货主编号*", "key": "shipper_id", "width": 16, "example": "SH001",
     "kind": "enum", "required": True, "label": "货主编号",
     "allowed": "SH001 / SH002", "fmt": "从下拉选择"},
    {"header": "客户编号*", "key": "customer_id", "width": 18, "example": "CUST001",
     "kind": "enum", "required": True, "label": "客户编号",
     "allowed": "CUST001 / CUST002 / CUST003", "fmt": "从下拉选择"},
    {"header": "目的地代码*", "key": "destination", "width": 18, "example": "CC",
     "kind": "enum", "required": True, "label": "目的地代码",
     "allowed": "CC / DL / TJ", "fmt": "CC=长春  DL=大连  TJ=天津"},
    {"header": "SKU编号*", "key": "sku_id", "width": 14, "example": "SKU001",
     "kind": "enum", "required": True, "label": "SKU编号",
     "allowed": "SKU001", "fmt": "从下拉选择"},
    {"header": "预测数量（托盘）*", "key": "quantity_pallets", "width": 22, "example": 15,
     "kind": "int", "required": True, "label": "预测数量",
     "allowed": "1–500 整数", "fmt": "托盘数，正整数"},
    {"header": "要求交付日期*", "key": "required_date", "width": 22, "example": "2026-06-23",
     "kind": "date", "required": True, "label": "要求交付日期", "snap": True,
     "allowed": "日期，YYYY-MM-DD",
     "fmt": "系统自动对齐至该周周一；例如填 2026-06-23（周二），存储为 2026-06-22（周一）"},
]

ORDER_COLUMNS = [
    {"header": "货主编号*", "key": "shipper_id", "width": 16, "example": "SH001",
     "kind": "enum", "required": True, "label": "货主编号",
     "allowed": "SH001 / SH002", "fmt": "从下拉选择"},
    {"header": "客户编号*", "key": "customer_id", "width": 18, "example": "CUST001",
     "kind": "enum", "required": True, "label": "客户编号",
     "allowed": "CUST001 / CUST002 / CUST003", "fmt": "从下拉选择"},
    {"header": "目的地代码*", "key": "destination", "width": 18, "example": "CC",
     "kind": "enum", "required": True, "label": "目的地代码",
     "allowed": "CC / DL / TJ", "fmt": "CC=长春  DL=大连  TJ=天津"},
    {"header": "SKU编号*", "key": "sku_id", "width": 14, "example": "SKU001",
     "kind": "enum", "required": True, "label": "SKU编号",
     "allowed": "SKU001", "fmt": "从下拉选择"},
    {"header": "确认数量（托盘）*", "key": "confirmed_quantity", "width": 22, "example": 10,
     "kind": "int", "required": True, "label": "确认数量",
     "allowed": "0–500 整数", "fmt": "托盘数，0 表示取消"},
    {"header": "确认交付日期*", "key": "confirmed_delivery_date", "width": 22, "example": "2026-06-23",
     "kind": "date", "required": True, "label": "确认交付日期", "snap": False,
     "allowed": "日期，YYYY-MM-DD", "fmt": "客户确认的实际交付日期，不做周对齐"},
    {"header": "客户订单号", "key": "order_no", "width": 22, "example": "PO-2026-00123",
     "kind": "text", "required": False, "label": "客户订单号",
     "allowed": "文本", "fmt": "选填，客户方采购订单号"},
]


def _build_template(title: str, sheet_name: str, columns: list) -> bytes:
    """Build an in-memory Excel workbook (data sheet + 填写说明) from a spec."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1F5C99", end_color="1F5C99", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    for idx, col in enumerate(columns, start=1):
        letter = openpyxl.utils.get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=col["header"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[letter].width = col["width"]

        # Example row
        ex = ws.cell(row=2, column=idx, value=col["example"])
        ex.alignment = Alignment(horizontal="center")

        # Per-column data validation
        dv = None
        if col["kind"] == "enum":
            values = _ENUMS[col["key"]]
            dv = DataValidation(
                type="list", formula1='"' + ",".join(values) + '"',
                allow_blank=not col["required"], showDropDown=False,
                showErrorMessage=True, errorTitle=f"无效{col['label']}",
                error="请从下拉列表选择：" + " / ".join(values),
            )
        elif col["kind"] == "int":
            dv = DataValidation(
                type="whole", operator="between", formula1="0", formula2="500",
                allow_blank=not col["required"], showErrorMessage=True,
                errorTitle="无效数量", error="请填写 0–500 之间的整数",
            )
        if dv is not None:
            ws.add_data_validation(dv)
            dv.sqref = f"{letter}2:{letter}1000"

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ── 填写说明 sheet ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("填写说明")
    title_font = Font(bold=True, size=13)
    ws2["A1"] = title
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")

    guide_headers = ["列", "字段名称", "是否必填", "允许值", "格式说明", "示例"]
    gh_font = Font(bold=True, size=10)
    gh_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, h in enumerate(guide_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = gh_font
        cell.fill = gh_fill
        cell.alignment = center

    for row_idx, col in enumerate(columns, start=4):
        letter = openpyxl.utils.get_column_letter(row_idx - 3)
        required_str = "是（*）" if col["required"] else "否"
        row_data = (letter, col["label"], required_str,
                    col["allowed"], col["fmt"], str(col["example"]))
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, w in enumerate([6, 20, 12, 30, 52, 16], start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 28
    for r in range(4, 4 + len(columns)):
        ws2.row_dimensions[r].height = 36

    note_row = len(columns) + 5
    ws2.cell(row=note_row, column=1, value="注意事项").font = Font(bold=True)
    notes = [
        "• 请勿修改第1行表头，否则上传时将无法识别列名",
        "• 第2行为示例数据，上传前请替换为实际数据（或直接在第3行起填写）",
        "• 日期列请使用 YYYY-MM-DD 格式，或 Excel 日期单元格格式均可",
        "• 上传后系统将展示解析预览，确认无误后再点击「确认提交」",
    ]
    for i, note in enumerate(notes, start=note_row + 1):
        ws2.cell(row=i, column=1, value=note)
        ws2.merge_cells(f"A{i}:F{i}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _validate(df_raw: pd.DataFrame, columns: list):
    """Parse and validate a raw uploaded DataFrame against a column spec.

    Pure function (no Streamlit calls) so it can be unit-tested.
    Returns ``(df, errors)``: cleaned/typed DataFrame when ``errors`` is
    empty, otherwise ``(None, errors)``.
    """
    col_map = {c["header"]: c["key"] for c in columns}
    inv = {c["key"]: c["header"] for c in columns}
    df = df_raw.rename(columns={h: col_map[h] for h in df_raw.columns if h in col_map})
    df = df.dropna(how="all").reset_index(drop=True)

    errors = []
    for col in columns:
        key = col["key"]
        if col["required"]:
            if key not in df.columns:
                errors.append(f"缺少必填列：**{inv[key]}**")
                continue
            if df[key].isna().any():
                bad = [r + 2 for r in df.index[df[key].isna()].tolist()]
                errors.append(f"列「{inv[key]}」第 {bad} 行存在空值")
        if col["kind"] == "enum" and key in df.columns:
            invalid = set(df[key].dropna().astype(str)) - set(_ENUMS[key])
            if invalid:
                errors.append(
                    f"无效{col['label']}：{invalid}，允许值为 {' / '.join(_ENUMS[key])}"
                )

    if errors:
        return None, errors

    # Clean / normalize
    for col in columns:
        key = col["key"]
        if key not in df.columns:
            continue
        if col["kind"] == "date":
            dt = pd.to_datetime(df[key])
            if col.get("snap"):
                df[key] = dt.apply(lambda d: (d - pd.Timedelta(days=d.weekday())).date())
            else:
                df[key] = dt.apply(lambda d: d.date())
        elif col["kind"] == "int":
            df[key] = df[key].astype(int)
        elif col["kind"] == "enum":
            df[key] = df[key].astype(str)
    return df, []


def _make_excel_template() -> bytes:
    """Forecast bulk-upload template (预测数据 + 填写说明)."""
    return _build_template("预测录入Excel模板 — 填写说明", "预测数据", FORECAST_COLUMNS)


def parse_and_validate(df_raw: pd.DataFrame):
    """Parse and validate an uploaded forecast DataFrame."""
    return _validate(df_raw, FORECAST_COLUMNS)


def _make_order_template() -> bytes:
    """Order-confirmation bulk-upload template (订单确认数据 + 填写说明)."""
    return _build_template("订单确认Excel模板 — 填写说明", "订单确认数据", ORDER_COLUMNS)


def parse_and_validate_order(df_raw: pd.DataFrame):
    """Parse and validate an uploaded order-confirmation DataFrame."""
    return _validate(df_raw, ORDER_COLUMNS)


# Columns that together define a forecast's identity for duplicate detection
FORECAST_STORE_COLS = ["shipper_id", "customer_id", "destination",
                       "sku_id", "quantity_pallets", "required_date"]


def _row_key(row) -> tuple:
    """Normalized identity tuple for a forecast row (NaN/None → '')."""
    out = []
    for c in FORECAST_STORE_COLS:
        v = row[c]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            v = ""
        out.append(v)
    return tuple(out)


def split_duplicates(existing: pd.DataFrame, incoming: pd.DataFrame):
    """Split ``incoming`` into ``(to_add, duplicates)``.

    A row is a duplicate if it matches an existing row OR an earlier
    incoming row on all FORECAST_STORE_COLS. Pure / testable.
    """
    seen = set()
    if existing is not None and len(existing):
        for _, r in existing.iterrows():
            seen.add(_row_key(r))

    add_rows, dup_rows = [], []
    for _, r in incoming.iterrows():
        k = _row_key(r)
        if k in seen:
            dup_rows.append(r)
        else:
            seen.add(k)
            add_rows.append(r)

    empty = incoming.iloc[0:0]
    to_add = pd.DataFrame(add_rows, columns=incoming.columns) if add_rows else empty
    dups = pd.DataFrame(dup_rows, columns=incoming.columns) if dup_rows else empty
    return to_add, dups


def _describe_forecast(row) -> str:
    """One-line human-readable description of a forecast row (for warnings)."""
    ship = row.get("shipper_id") if hasattr(row, "get") else row["shipper_id"]
    ship = "" if (ship is None or (isinstance(ship, float) and pd.isna(ship))) else ship
    return (f"货主 {ship or '—'} / 客户 {row['customer_id']} / 目的地 {row['destination']} "
            f"/ {row['sku_id']} / {int(row['quantity_pallets'])}托盘 / {row['required_date']}")


st.set_page_config(page_title="销售预测管理", page_icon="📈", layout="wide")
st.title("📈 销售预测管理")
st.caption("预测录入、置信度修正、订单确认")

# In-session forecast store (no DB yet); seeded with a few demo rows
if "forecasts" not in st.session_state:
    st.session_state.forecasts = pd.DataFrame(
        [
            {"shipper_id": "SH001", "customer_id": "CUST001", "destination": "CC",
             "sku_id": "SKU001", "quantity_pallets": 15, "required_date": date(2026, 4, 27)},
            {"shipper_id": "SH001", "customer_id": "CUST002", "destination": "DL",
             "sku_id": "SKU001", "quantity_pallets": 8, "required_date": date(2026, 4, 27)},
            {"shipper_id": "SH002", "customer_id": "CUST003", "destination": "TJ",
             "sku_id": "SKU001", "quantity_pallets": 20, "required_date": date(2026, 5, 4)},
        ],
        columns=FORECAST_STORE_COLS,
    )

tab1, tab2, tab3 = st.tabs(["预测录入", "预测列表", "订单确认"])

with tab1:
    st.subheader("新增销售预测")

    with st.form("new_forecast"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            customer = st.selectbox("客户", CUSTOMERS)
        with col2:
            dest_label = st.selectbox("目的地", list(DEST_OPTIONS.keys()))
        with col3:
            sku = st.selectbox("SKU", SKU_OPTIONS)
        with col4:
            qty = st.number_input("预测数量(托盘)", min_value=1, max_value=100, value=10)

        today = date.today()
        wcol1, wcol2 = st.columns(2)
        with wcol1:
            year = st.number_input("要求交付年份", min_value=2025, max_value=2030,
                                   value=today.year, step=1)
        with wcol2:
            max_week = date(int(year), 12, 28).isocalendar()[1]
            cur_week = today.isocalendar()[1]
            week_num = st.number_input("要求交付周（ISO周）", min_value=1,
                                       max_value=int(max_week),
                                       value=int(cur_week), step=1)
        due = date.fromisocalendar(int(year), int(week_num), 1)

        submitted = st.form_submit_button("提交预测")
        if submitted:
            new_row = pd.DataFrame([{
                "shipper_id": "",
                "customer_id": customer,
                "destination": DEST_OPTIONS[dest_label],
                "sku_id": sku,
                "quantity_pallets": int(qty),
                "required_date": due,
            }], columns=FORECAST_STORE_COLS)
            to_add, dups = split_duplicates(st.session_state.forecasts, new_row)
            if len(dups):
                st.warning(
                    "⚠️ 该预测信息与已有记录完全重复："
                    f"{_describe_forecast(new_row.iloc[0])}"
                )
            else:
                st.session_state.forecasts = pd.concat(
                    [st.session_state.forecasts, to_add], ignore_index=True)
                st.success(
                    f"预测已提交: {sku} {qty}托盘 → {dest_label}, 第 {week_num} 周")

    st.divider()
    st.markdown("#### 批量导入")

    bcol1, bcol2 = st.columns([1, 3])
    with bcol1:
        st.download_button(
            label="📥 下载Excel模板",
            data=_make_excel_template(),
            file_name="预测录入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with bcol2:
        uploaded = st.file_uploader(
            "上传填好的Excel文件（仅支持 .xlsx）",
            type=["xlsx"],
        )

    if uploaded is not None:
        try:
            df_raw = pd.read_excel(uploaded, sheet_name="预测数据", header=0)
        except Exception as e:
            st.error(f"无法读取文件：{e}，请确认上传的是从模板下载的 .xlsx 文件")
            st.stop()

        df, errors = parse_and_validate(df_raw)

        if errors:
            st.error("上传文件存在以下问题，请修正后重新上传：")
            for e in errors:
                st.markdown(f"- {e}")
        else:
            display_cols = [c for c in
                ["shipper_id", "customer_id", "destination", "sku_id",
                 "quantity_pallets", "required_date"]
                if c in df.columns]

            st.success(f"解析成功，共 {len(df)} 条预测，请核对后提交")
            st.dataframe(
                df[display_cols].rename(columns={
                    "shipper_id": "货主编号",
                    "customer_id": "客户编号",
                    "destination": "目的地",
                    "sku_id": "SKU",
                    "quantity_pallets": "预测数量（托盘）",
                    "required_date": "要求日期（周一）",
                }),
                use_container_width=True,
                hide_index=True,
            )

            if st.button("✅ 确认提交全部预测", type="primary"):
                to_add, dups = split_duplicates(
                    st.session_state.forecasts, df[FORECAST_STORE_COLS])
                st.session_state.forecasts = pd.concat(
                    [st.session_state.forecasts, to_add], ignore_index=True)
                if len(dups):
                    st.warning(f"⚠️ 以下 {len(dups)} 条预测信息完全重复，已自动跳过：")
                    for _, r in dups.iterrows():
                        st.markdown(f"- {_describe_forecast(r)}")
                st.success(
                    f"已提交 {len(to_add)} 条预测"
                    f"（重复 {len(dups)} 条已跳过；Phase 2 接入引擎后将写入数据库）")

with tab2:
    st.subheader("已提交预测列表")
    st.caption("完全重复的预测信息不会重复显示")

    fc = st.session_state.forecasts.drop_duplicates(
        subset=FORECAST_STORE_COLS).reset_index(drop=True)

    if fc.empty:
        st.info("暂无已提交预测")
    else:
        st.dataframe(
            fc.rename(columns={
                "shipper_id": "货主",
                "customer_id": "客户",
                "destination": "目的地",
                "sku_id": "SKU",
                "quantity_pallets": "预测量",
                "required_date": "要求日期",
            }),
            use_container_width=True, hide_index=True,
        )

with tab3:
    st.subheader("订单确认")
    st.caption("提交客户订单确认，后台自动匹配对应预测（无预测记录时也可提交）")

    with st.form("confirm_order"):
        col1, col2, col3 = st.columns(3)
        with col1:
            c_customer = st.selectbox("客户", CUSTOMERS)
        with col2:
            c_dest_label = st.selectbox("目的地", list(DEST_OPTIONS.keys()))
        with col3:
            c_sku = st.selectbox("SKU", SKU_OPTIONS)

        col4, col5 = st.columns(2)
        with col4:
            c_date = st.date_input("确认交付日期", value=date.today())
        with col5:
            c_qty = st.number_input("确认数量(托盘)", min_value=0, max_value=500, value=10)

        c_order_no = st.text_input("客户订单号", placeholder="如 PO-2026-00123")

        if st.form_submit_button("提交确认"):
            st.success(
                f"订单已提交：{c_customer} → {c_dest_label}，{c_sku}，"
                f"{c_date}，{c_qty} 托盘，订单号 {c_order_no}"
            )

    st.divider()
    st.markdown("#### 批量导入")

    ocol1, ocol2 = st.columns([1, 3])
    with ocol1:
        st.download_button(
            label="📥 下载Excel模板",
            data=_make_order_template(),
            file_name="订单确认模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="order_template_dl",
        )
    with ocol2:
        order_uploaded = st.file_uploader(
            "上传填好的Excel文件（仅支持 .xlsx）",
            type=["xlsx"],
            key="order_upload",
        )

    if order_uploaded is not None:
        try:
            odf_raw = pd.read_excel(order_uploaded, sheet_name="订单确认数据", header=0)
        except Exception as e:
            st.error(f"无法读取文件：{e}，请确认上传的是从模板下载的 .xlsx 文件")
            st.stop()

        odf, oerrors = parse_and_validate_order(odf_raw)

        if oerrors:
            st.error("上传文件存在以下问题，请修正后重新上传：")
            for e in oerrors:
                st.markdown(f"- {e}")
        else:
            display_cols = [c["key"] for c in ORDER_COLUMNS if c["key"] in odf.columns]
            zh_names = {c["key"]: c["header"].rstrip("*") for c in ORDER_COLUMNS}

            st.success(f"解析成功，共 {len(odf)} 条订单确认，请核对后提交")
            st.dataframe(
                odf[display_cols].rename(columns=zh_names),
                use_container_width=True,
                hide_index=True,
            )

            if st.button("✅ 确认提交全部订单", type="primary", key="order_submit"):
                st.success(f"已提交 {len(odf)} 条订单确认（Phase 5 接入引擎后将写入数据库）")
